"""
Экспорт аналитических данных в Excel.

Конкурсное требование:
  «Функционал выгрузки аналитической информации из программы»
  - статистика по посещаемости ТТ
  - отчёт о времени нахождения торгового представителя на каждой ТТ
  - детализация по времени и дате посещения каждой точки
  - учёт количества выходов торгового представителя на маршрут
"""

import io
import logging
from calendar import monthrange
from collections import defaultdict
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from src.database.models import Location, SalesRep, VisitLog, VisitSchedule, get_session

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

router = APIRouter(prefix="/export", tags=["Export"])
logger = logging.getLogger("export")

# ─── Цвета заголовков по категориям ─────────────────────────────────────────
CAT_COLORS = {
    "A": "FFEF4444",  # красный
    "B": "FFF97316",  # оранжевый
    "C": "FFEAB308",  # жёлтый
    "D": "FF6B7280",  # серый
}
HEADER_FILL = PatternFill("solid", fgColor="FF1D4ED8") if XLSX_AVAILABLE else None
HEADER_FONT = Font(color="FFFFFFFF", bold=True) if XLSX_AVAILABLE else None


def _apply_header(ws, row: int, cols: list[str]) -> None:
    """Записывает строку заголовков с форматированием."""
    for c, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20


def _autofit(ws) -> None:
    """Авто-ширина столбцов (по максимальной длине значения)."""
    col_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                col_widths[cell.column] = max(
                    col_widths.get(cell.column, 0),
                    len(str(cell.value)),
                )
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(w + 2, 50)


# ─── Главный эндпоинт ────────────────────────────────────────────────────────

@router.get("/schedule")
async def export_schedule(
    month: str = Query(..., description="Месяц YYYY-MM"),
    session: AsyncSession = Depends(get_session),
):
    """
    Выгружает план расписания и аналитику визитов в Excel.

    Листы:
      1. Расписание        — все плановые визиты
      2. Журнал визитов    — фактические визиты с временем
      3. Статистика по ТТ — охват, выполнение по категориям
      4. Активность ТП    — количество выходов и посещённых ТТ
    """
    if not XLSX_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl не установлен")

    try:
        year, m = map(int, month.split("-"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Формат месяца: YYYY-MM")

    _, last_day = monthrange(year, m)
    month_start = date(year, m, 1)
    month_end = date(year, m, last_day)

    # ── Загружаем данные ─────────────────────────────────────────────────────
    sched_q = (
        select(VisitSchedule)
        .where(
            VisitSchedule.planned_date >= month_start,
            VisitSchedule.planned_date <= month_end,
        )
        .options(
            selectinload(VisitSchedule.location),
            selectinload(VisitSchedule.rep),
        )
    )
    schedules = (await session.execute(sched_q)).scalars().all()

    log_q = (
        select(VisitLog)
        .where(
            VisitLog.visited_date >= month_start,
            VisitLog.visited_date <= month_end,
        )
    )
    logs = (await session.execute(log_q)).scalars().all()
    logs_by_sched: dict[str, VisitLog] = {lg.schedule_id: lg for lg in logs}

    all_locs = (await session.execute(select(Location))).scalars().all()
    all_reps = (await session.execute(select(SalesRep))).scalars().all()
    rep_map = {r.id: r.name for r in all_reps}
    loc_map = {loc.id: loc for loc in all_locs}

    # ── Создаём книгу ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # Лист 1: Расписание
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Расписание"
    ws1.merge_cells("A1:H1")
    ws1["A1"] = f"ПЛАН ПОСЕЩЕНИЙ — {month}"
    ws1["A1"].font = Font(bold=True, size=13)
    ws1["A1"].alignment = Alignment(horizontal="center")

    SCHED_COLS = [
        "Дата", "Сотрудник", "ТТ", "Адрес",
        "Категория", "Статус", "Время прихода", "Время ухода",
    ]
    _apply_header(ws1, 2, SCHED_COLS)

    for i, s in enumerate(
        sorted(schedules, key=lambda x: (x.planned_date, x.rep_id)), start=3
    ):
        lg = logs_by_sched.get(s.id)
        loc = s.location
        cat = loc.category if loc else "?"

        # Цвет строки по категории
        fill_color = CAT_COLORS.get(cat or "?", "FFFFFFFF")

        row_data = [
            s.planned_date.strftime("%d.%m.%Y") if s.planned_date else "",
            rep_map.get(s.rep_id, s.rep_id),
            loc.name if loc else s.location_id,
            getattr(loc, "address", "") or "",
            cat,
            _status_label(s.status),
            str(lg.time_in)[:5] if lg and lg.time_in else "",
            str(lg.time_out)[:5] if lg and lg.time_out else "",
        ]
        for c, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=i, column=c, value=val)
            if cat and cat in CAT_COLORS and c == 5:
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(bold=True, color="FFFFFFFF")

    _autofit(ws1)

    # ════════════════════════════════════════════════════════════════════════
    # Лист 2: Журнал визитов (только выполненные)
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Журнал визитов")
    ws2.merge_cells("A1:G1")
    ws2["A1"] = f"ЖУРНАЛ ФАКТИЧЕСКИХ ВИЗИТОВ — {month}"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A1"].alignment = Alignment(horizontal="center")

    LOG_COLS = [
        "Дата визита", "Сотрудник", "ТТ", "Категория",
        "Время прихода", "Время ухода", "Длительность (мин)",
    ]
    _apply_header(ws2, 2, LOG_COLS)

    completed = [s for s in schedules if s.status == "completed"]
    for i, s in enumerate(
        sorted(completed, key=lambda x: (x.planned_date, x.rep_id)), start=3
    ):
        lg = logs_by_sched.get(s.id)
        loc = s.location
        duration = _calc_duration(lg)
        ws2.append([])  # append нельзя с row-gap, используем cell
        row_data = [
            s.planned_date.strftime("%d.%m.%Y") if s.planned_date else "",
            rep_map.get(s.rep_id, s.rep_id),
            loc.name if loc else s.location_id,
            loc.category if loc else "?",
            str(lg.time_in)[:5] if lg and lg.time_in else "—",
            str(lg.time_out)[:5] if lg and lg.time_out else "—",
            duration,
        ]
        for c, val in enumerate(row_data, start=1):
            ws2.cell(row=i, column=c, value=val)

    _autofit(ws2)

    # ════════════════════════════════════════════════════════════════════════
    # Лист 3: Статистика по ТТ
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Статистика по ТТ")
    ws3.merge_cells("A1:F1")
    ws3["A1"] = f"ОХВАТ И СТАТИСТИКА ТОРГОВЫХ ТОЧЕК — {month}"
    ws3["A1"].font = Font(bold=True, size=13)
    ws3["A1"].alignment = Alignment(horizontal="center")

    STAT_COLS = [
        "ТТ", "Категория", "Район", "Запланировано",
        "Выполнено", "Пропущено", "% выполнения",
    ]
    _apply_header(ws3, 2, STAT_COLS)

    # Агрегируем по ТТ
    tt_stats: dict[str, dict] = {}
    for s in schedules:
        loc_id = s.location_id
        if loc_id not in tt_stats:
            loc = s.location or loc_map.get(loc_id)
            tt_stats[loc_id] = {
                "name": loc.name if loc else loc_id,
                "category": loc.category if loc else "?",
                "district": getattr(loc, "district", "") or "",
                "planned": 0, "completed": 0, "skipped": 0,
            }
        tt_stats[loc_id]["planned"] += 1
        if s.status == "completed":
            tt_stats[loc_id]["completed"] += 1
        elif s.status == "skipped":
            tt_stats[loc_id]["skipped"] += 1

    for i, (loc_id, st) in enumerate(
        sorted(tt_stats.items(), key=lambda x: (x[1]["category"], x[1]["name"])),
        start=3,
    ):
        pct = round(st["completed"] / st["planned"] * 100, 1) if st["planned"] else 0
        ws3.cell(row=i, column=1, value=st["name"])
        ws3.cell(row=i, column=2, value=st["category"])
        ws3.cell(row=i, column=3, value=st["district"])
        ws3.cell(row=i, column=4, value=st["planned"])
        ws3.cell(row=i, column=5, value=st["completed"])
        ws3.cell(row=i, column=6, value=st["skipped"])
        pct_cell = ws3.cell(row=i, column=7, value=f"{pct}%")
        if pct >= 80:
            pct_cell.font = Font(color="FF16A34A", bold=True)
        elif pct < 50:
            pct_cell.font = Font(color="FFDC2626", bold=True)

    _autofit(ws3)

    # ════════════════════════════════════════════════════════════════════════
    # Лист 4: Активность торговых представителей
    # ════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Активность ТП")
    ws4.merge_cells("A1:F1")
    ws4["A1"] = f"АКТИВНОСТЬ ТОРГОВЫХ ПРЕДСТАВИТЕЛЕЙ — {month}"
    ws4["A1"].font = Font(bold=True, size=13)
    ws4["A1"].alignment = Alignment(horizontal="center")

    ACT_COLS = [
        "Сотрудник", "Выходов на маршрут", "ТТ запланировано",
        "ТТ выполнено", "ТТ пропущено", "% выполнения",
    ]
    _apply_header(ws4, 2, ACT_COLS)

    rep_stats: dict[str, dict] = {}
    for s in schedules:
        rid = s.rep_id
        if rid not in rep_stats:
            rep_stats[rid] = {
                "name": rep_map.get(rid, rid),
                "outings": set(),
                "planned": 0, "completed": 0, "skipped": 0,
            }
        if s.planned_date:
            rep_stats[rid]["outings"].add(s.planned_date)
        rep_stats[rid]["planned"] += 1
        if s.status == "completed":
            rep_stats[rid]["completed"] += 1
        elif s.status == "skipped":
            rep_stats[rid]["skipped"] += 1

    for i, (rid, rs) in enumerate(
        sorted(rep_stats.items(), key=lambda x: x[1]["name"]),
        start=3,
    ):
        pct = round(rs["completed"] / rs["planned"] * 100, 1) if rs["planned"] else 0
        ws4.cell(row=i, column=1, value=rs["name"])
        ws4.cell(row=i, column=2, value=len(rs["outings"]))  # уникальные дни с визитами
        ws4.cell(row=i, column=3, value=rs["planned"])
        ws4.cell(row=i, column=4, value=rs["completed"])
        ws4.cell(row=i, column=5, value=rs["skipped"])
        ws4.cell(row=i, column=6, value=f"{pct}%")

    _autofit(ws4)

    # ── Отдаём файл ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"t2_schedule_{month}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# ─── Хелперы ─────────────────────────────────────────────────────────────────

def _status_label(status: str) -> str:
    return {
        "planned":     "Запланирован",
        "completed":   "Выполнен",
        "skipped":     "Пропущен",
        "cancelled":   "Отменён",
        "rescheduled": "Перенесён",
    }.get(status, status)


def _calc_duration(log) -> str:
    """Возвращает длительность визита в минутах или пустую строку."""
    if not log or not log.time_in or not log.time_out:
        return ""
    try:
        t_in = log.time_in
        t_out = log.time_out
        minutes = (t_out.hour * 60 + t_out.minute) - (t_in.hour * 60 + t_in.minute)
        return str(minutes) if minutes > 0 else ""
    except Exception:
        return ""
