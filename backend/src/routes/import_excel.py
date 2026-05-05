"""
Импорт данных визитов из заполненного Excel-отчёта.

POST /api/v1/import/schedule
  — читает лист «Расписание» из файла, ранее экспортированного /export/schedule
  — обновляет статусы VisitSchedule и создаёт/обновляет записи VisitLog
  — возвращает { updated, skipped, errors }
"""

import io
import logging
from datetime import date, time as dtime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.database.models import Location, SalesRep, VisitLog, VisitSchedule, get_session

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

router = APIRouter(prefix="/import", tags=["Import"])
logger = logging.getLogger("import_excel")

STATUS_MAP = {
    "Выполнен":     "completed",
    "Запланирован": "planned",
    "Пропущен":     "skipped",
    "Отменён":      "cancelled",
    "Перенесён":    "rescheduled",
}

HEADER_ALIASES = {
    "date": {"дата"},
    "rep_name": {"сотрудник"},
    "location_name": {"тт"},
    "status": {"статус"},
    "time_in": {"время прихода"},
    "time_out": {"время ухода"},
    "schedule_id": {"id визита"},
    "rep_id": {"id сотрудника"},
    "location_id": {"id тт"},
}


def _parse_time(raw) -> dtime | None:
    if not raw:
        return None
    s = str(raw).strip()
    if not s or s == "—":
        return None
    try:
        parts = s.split(":")
        return dtime(int(parts[0]), int(parts[1]))
    except Exception:
        return None


def _parse_date(raw) -> date | None:
    if not raw:
        return None
    s = str(raw).strip()
    try:
        if "." in s:
            d, m, y = s.split(".")
            return date(int(y), int(m), int(d))
        # Если openpyxl вернул datetime-объект
        if hasattr(raw, "date"):
            return raw.date()
    except Exception:
        pass
    return None


def _normalize_header(raw) -> str:
    return str(raw or "").strip().lower()


def _build_header_index(header_row) -> dict[str, int]:
    index: dict[str, int] = {}
    for idx, raw_header in enumerate(header_row):
        normalized = _normalize_header(raw_header)
        if not normalized:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                index[field] = idx
                break
    return index


def _get_row_value(row, header_index: dict[str, int], field: str):
    idx = header_index.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


@router.post("/schedule")
async def import_schedule_excel(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
):
    """
    Загружает заполненный Excel-файл с результатами визитов.

    Формат файла: экспорт /export/schedule (лист «Расписание»).
    Заполненные поля «Статус», «Время прихода», «Время ухода»
    обновляют базу данных и пополняют аналитику.

    Возвращает: { updated, skipped, errors }.
    """
    if not XLSX_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl не установлен")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать Excel: {exc}")

    ws = None
    if "Расписание" in wb.sheetnames:
        ws = wb["Расписание"]
    elif wb.active:
        ws = wb.active

    if ws is None:
        raise HTTPException(status_code=400, detail="Лист 'Расписание' не найден")

    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
    header_index = _build_header_index(header_row or [])
    if "date" not in header_index or "status" not in header_index:
        raise HTTPException(
            status_code=400,
            detail="В Excel не найдены обязательные колонки 'Дата' и 'Статус'",
        )

    # Загружаем словари для матчинга по имени
    rep_by_name = {
        r.name: r
        for r in (await session.execute(select(SalesRep))).scalars().all()
    }
    loc_by_name = {
        loc.name: loc
        for loc in (await session.execute(select(Location))).scalars().all()
    }

    updated = 0
    skipped = 0
    errors: list[str] = []

    # Строки начинаются с 3 (1 — заголовок отчёта, 2 — шапка таблицы)
    for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or row[0] is None:
            continue

        # Читаем значения колонок
        raw_date = _get_row_value(row, header_index, "date")
        rep_name = str(_get_row_value(row, header_index, "rep_name") or "").strip()
        loc_name = str(_get_row_value(row, header_index, "location_name") or "").strip()
        status_ru = str(_get_row_value(row, header_index, "status") or "").strip()
        raw_tin = _get_row_value(row, header_index, "time_in")
        raw_tout = _get_row_value(row, header_index, "time_out")
        schedule_id = str(_get_row_value(row, header_index, "schedule_id") or "").strip()
        rep_id = str(_get_row_value(row, header_index, "rep_id") or "").strip()
        location_id = str(_get_row_value(row, header_index, "location_id") or "").strip()

        planned_date = _parse_date(raw_date)
        if not planned_date:
            skipped += 1
            errors.append(f"Стр.{row_num}: неверный формат даты '{raw_date}'")
            continue

        internal_status = STATUS_MAP.get(status_ru)
        if not internal_status:
            skipped += 1
            errors.append(f"Стр.{row_num}: неизвестный статус '{status_ru}'")
            continue

        sched = None
        rep = None
        loc = None

        if schedule_id:
            sched = await session.get(VisitSchedule, schedule_id)

        if sched is None and rep_id and location_id:
            sched_q = select(VisitSchedule).where(
                VisitSchedule.planned_date == planned_date,
                VisitSchedule.rep_id == rep_id,
                VisitSchedule.location_id == location_id,
            )
            sched = (await session.execute(sched_q)).scalar_one_or_none()

        if sched is None:
            rep = rep_by_name.get(rep_name)
            if not rep:
                skipped += 1
                errors.append(f"Стр.{row_num}: сотрудник '{rep_name}' не найден")
                continue

            loc = loc_by_name.get(loc_name)
            if not loc:
                skipped += 1
                errors.append(f"Стр.{row_num}: ТТ '{loc_name}' не найдена")
                continue

            sched_q = select(VisitSchedule).where(
                VisitSchedule.planned_date == planned_date,
                VisitSchedule.rep_id == rep.id,
                VisitSchedule.location_id == loc.id,
            )
            sched = (await session.execute(sched_q)).scalar_one_or_none()

        if not sched:
            skipped += 1
            errors.append(
                f"Стр.{row_num}: визит не найден в расписании "
                f"({planned_date}, {rep_name}, {loc_name})"
            )
            continue

        if rep is None:
            rep = await session.get(SalesRep, sched.rep_id)
        if loc is None:
            loc = await session.get(Location, sched.location_id)

        # Administrative import: bypass VALID_TRANSITIONS intentionally (Excel data import = override)
        sched.status = internal_status

        # Создаём/обновляем VisitLog если визит выполнен и есть время
        if internal_status == "completed":
            t_in  = _parse_time(raw_tin)
            t_out = _parse_time(raw_tout)

            if t_in and t_out:
                log_q = select(VisitLog).where(VisitLog.schedule_id == sched.id)
                existing = (await session.execute(log_q)).scalar_one_or_none()

                if existing:
                    existing.time_in  = t_in
                    existing.time_out = t_out
                else:
                    session.add(VisitLog(
                        schedule_id=sched.id,
                        rep_id=rep.id,
                        location_id=loc.id,
                        visited_date=planned_date,
                        time_in=t_in,
                        time_out=t_out,
                    ))

        updated += 1

    try:
        await session.commit()
    except Exception as exc:
        await session.rollback()
        logger.error("import_schedule_excel commit failed: %s", exc)
        raise HTTPException(status_code=500, detail=f"Ошибка сохранения в БД: {exc}")

    logger.info("Excel import done: updated=%d skipped=%d", updated, skipped)
    return {
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20],  # не возвращаем тысячи строк
    }
