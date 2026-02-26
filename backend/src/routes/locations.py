import csv
import io
import json
from typing import List

from openpyxl import load_workbook

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.database.models import Location, get_session
from src.schemas.locations import (
    LocationCreate,
    LocationResponse,
    UploadLocationsResponse,
)


router = APIRouter(prefix="/locations", tags=["Locations"])


@router.get(
    "/",
    response_model=List[LocationResponse],
    status_code=status.HTTP_200_OK,
    responses={
        200: {"description": "Успешное получение списка всех локаций"},
        500: {"description": "Internal Server Error"},
    },
)
async def get_locations(
    session: AsyncSession = Depends(get_session),
):
    """
    Get all locations from the database.
    
    Returns a list of locations with their IDs, coordinates, and time windows.
    """
    # Создаем запрос на выборку всех записей из таблицы Location
    stmt = select(Location)
    result = await session.execute(stmt)
    
    # scalars() извлекает объекты моделей, а all() собирает их в список
    locations = result.scalars().all()
    
    return locations


@router.post(
    "/",
    response_model=LocationResponse,
    status_code=status.HTTP_201_CREATED,
    responses={
        400: {"description": "Bad Request"},
        422: {"description": "Validation Error"},
        500: {"description": "Internal Server Error"},
    },
)
async def create_location(
    location_data: LocationCreate,
    session: AsyncSession = Depends(get_session),
):
    """Create a new location in the database."""
    new_location = Location(**location_data.model_dump())
    session.add(new_location)
    await session.commit()
    await session.refresh(new_location)

    return new_location


@router.post(
    "/upload",
    response_model=UploadLocationsResponse,
    status_code=status.HTTP_201_CREATED,
    responses={
        400: {"description": "Unsupported file format"},
        422: {"description": "Validation Error"},
        500: {"description": "Internal Server Error"},
    },
)
async def upload_locations(
    file: UploadFile,
    session: AsyncSession = Depends(get_session),
):
    """Upload locations from a CSV, JSON, or XLSX file.

    CSV format: name,lat,lon,time_window_start,time_window_end
    JSON format: array of objects with the same fields.
    XLSX format: first row = headers, columns matching CSV fields.
    """
    filename = (file.filename or "").lower()
    content = await file.read()

    if filename.endswith(".json"):
        rows = _parse_json(content)
    elif filename.endswith(".csv"):
        rows = _parse_csv(content)
    elif filename.endswith(".xlsx"):
        rows = _parse_xlsx(content)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file format. Use .csv, .json, or .xlsx",
        )

    created: list[Location] = []
    errors: list[dict] = []

    for idx, row in enumerate(rows):
        try:
            loc_data = LocationCreate(**row)
            new_location = Location(**loc_data.model_dump())
            session.add(new_location)
            await session.flush()
            created.append(new_location)
        except Exception as exc:
            errors.append({"row": idx + 1, "error": str(exc), "data": row})

    if created:
        await session.commit()
        for loc in created:
            await session.refresh(loc)

    return UploadLocationsResponse(
        created=[LocationResponse.model_validate(loc) for loc in created],
        errors=errors,
        total_processed=len(rows),
    )


def _parse_json(content: bytes) -> list[dict]:
    """Parse JSON file content into list of dicts."""
    data = json.loads(content.decode("utf-8"))
    if isinstance(data, list):
        return data
    raise ValueError("JSON file must contain an array of objects")


def _parse_csv(content: bytes) -> list[dict]:
    """Parse CSV file content into list of dicts."""
    text = content.decode("utf-8")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        parsed = {}
        for key, value in row.items():
            key = key.strip()
            value = value.strip() if value else value
            if key in ("lat", "lon"):
                parsed[key] = float(value)
            else:
                parsed[key] = value
        rows.append(parsed)
    return rows


# Column name aliases for XLSX files (Russian -> English)
_XLSX_COLUMN_MAP = {
    "name": "name",
    "название": "name",
    "наименование": "name",
    "наименование точки": "name",
    "наименование торговой точки": "name",
    "наименование района": "name",       # формат тестового файла организаторов
    "район": "name",
    "торговая точка": "name",
    "тт": "name",
    "адрес": "name",
    "lat": "lat",
    "latitude": "lat",
    "широта": "lat",
    "lon": "lon",
    "lng": "lon",
    "longitude": "lon",
    "долгота": "lon",
    "time_window_start": "time_window_start",
    "начало": "time_window_start",
    "время начала": "time_window_start",
    "открытие": "time_window_start",
    "time_window_end": "time_window_end",
    "конец": "time_window_end",
    "время окончания": "time_window_end",
    "закрытие": "time_window_end",
    "category": "category",
    "категория": "category",
    "кат": "category",
    "city": "city",
    "город": "city",
    "district": "district",
    "район": "district",
    "address": "address",
    "адрес тт": "address",
}

# Справочник координат районов и городов Республики Мордовия.
# Используется когда в файле есть название района, но нет координат.
_MORDOVIA_COORDS: dict[str, tuple[float, float]] = {
    # Городской округ
    "г.о. саранск":         (54.1871, 45.1749),
    "саранск":              (54.1871, 45.1749),
    # Районы
    "ардатовский":          (54.8490, 46.2360),
    "ардатовский район":    (54.8490, 46.2360),
    "атюрьевский":          (54.0310, 43.6820),
    "атюрьевский район":    (54.0310, 43.6820),
    "атяшевский":           (54.5980, 45.8880),
    "атяшевский район":     (54.5980, 45.8880),
    "большеберезниковский": (54.2670, 45.7650),
    "большеберезниковский район": (54.2670, 45.7650),
    "большеигнатовский":    (54.4810, 44.8710),
    "большеигнатовский район": (54.4810, 44.8710),
    "дубёнский":            (54.2650, 46.0880),
    "дубенский":            (54.2650, 46.0880),
    "дубёнский район":      (54.2650, 46.0880),
    "дубенский район":      (54.2650, 46.0880),
    "ельниковский":         (54.3900, 43.5550),
    "ельниковский район":   (54.3900, 43.5550),
    "зубово-полянский":     (54.0520, 42.8310),
    "зубово-полянский район": (54.0520, 42.8310),
    "инсарский":            (53.8760, 44.3770),
    "инсарский район":      (53.8760, 44.3770),
    "ичалковский":          (54.1260, 46.5840),
    "ичалковский район":    (54.1260, 46.5840),
    "кадошкинский":         (54.0200, 43.5360),
    "кадошкинский район":   (54.0200, 43.5360),
    "ковылкинский":         (53.9060, 43.9190),
    "ковылкинский район":   (53.9060, 43.9190),
    "кочкуровский":         (54.0500, 45.5710),
    "кочкуровский район":   (54.0500, 45.5710),
    "краснослободский":     (54.4190, 43.7770),
    "краснослободский район": (54.4190, 43.7770),
    "лямбирский":           (54.2350, 45.6250),
    "лямбирский район":     (54.2350, 45.6250),
    "ромодановский":        (54.4300, 45.3710),
    "ромодановский район":  (54.4300, 45.3710),
    "рузаевский":           (54.0570, 44.9520),
    "рузаевский район":     (54.0570, 44.9520),
    "старошайговский":      (54.3500, 44.2580),
    "старошайговский район": (54.3500, 44.2580),
    "темниковский":         (54.6360, 43.1990),
    "темниковский район":   (54.6360, 43.1990),
    "теньгушевский":        (54.8440, 43.6140),
    "теньгушевский район":  (54.8440, 43.6140),
    "торбеевский":          (54.0880, 43.0920),
    "торбеевский район":    (54.0880, 43.0920),
    "чамзинский":           (54.2310, 46.2890),
    "чамзинский район":     (54.2310, 46.2890),
}

# Центр Саранска — запасные координаты если район не найден в справочнике
_DEFAULT_LAT = 54.1871
_DEFAULT_LON = 45.1749


def _lookup_coords(name: str) -> tuple[float, float] | None:
    """Ищет координаты по нормализованному названию района."""
    key = name.strip().lower()
    if key in _MORDOVIA_COORDS:
        return _MORDOVIA_COORDS[key]
    # Попытка частичного совпадения (без слова "район")
    for k, v in _MORDOVIA_COORDS.items():
        if key in k or k in key:
            return v
    return None


def _parse_xlsx(content: bytes) -> list[dict]:
    """Parse XLSX file content into list of dicts.

    Читает первый лист, первая строка = заголовки.
    Поддерживает русские псевдонимы колонок.
    Если координаты отсутствуют — ищет в справочнике районов Мордовии,
    при отсутствии совпадения — расставляет точки вокруг центра Саранска.
    """
    import random
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    # Берём первый лист по индексу, не wb.active (может быть неверным)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)

    raw_headers = next(rows_iter, None)
    if not raw_headers:
        wb.close()
        return []

    headers = []
    for h in raw_headers:
        h_str = str(h).strip().lower() if h else ""
        headers.append(_XLSX_COLUMN_MAP.get(h_str, h_str))

    rows = []
    for row_values in rows_iter:
        if all(v is None for v in row_values):
            continue

        parsed: dict = {}
        for header, value in zip(headers, row_values):
            if value is None:
                continue
            if header in ("lat", "lon"):
                try:
                    parsed[header] = float(value)
                except (ValueError, TypeError):
                    pass
            else:
                str_val = str(value).strip()
                # Пропускаем Excel-формулы
                if str_val.startswith("="):
                    continue
                if str_val:
                    parsed[header] = str_val

        if "name" not in parsed or not parsed["name"]:
            continue

        # Координаты: сначала справочник, потом случайный разброс
        if "lat" not in parsed or "lon" not in parsed:
            coords = _lookup_coords(parsed["name"])
            if coords:
                parsed["lat"], parsed["lon"] = coords
            else:
                parsed["lat"] = round(
                    _DEFAULT_LAT + random.uniform(-0.02, 0.02), 6
                )
                parsed["lon"] = round(
                    _DEFAULT_LON + random.uniform(-0.02, 0.02), 6
                )

        parsed.setdefault("time_window_start", "09:00")
        parsed.setdefault("time_window_end", "18:00")
        rows.append(parsed)

    wb.close()
    return rows